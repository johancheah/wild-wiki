from __future__ import annotations

import argparse
import logging
import re
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl

# Python's fromisoformat (pre-3.11) only accepts 3- or 6-digit fractional
# seconds — the API has been observed emitting other lengths (e.g. 1 digit,
# confirmed 2026-08-31). Pad/truncate to exactly 6 (microseconds) so any
# length parses — same fix as queries.py::_local_date.
_FRACTIONAL_SECONDS_RE = re.compile(r"\.(\d+)")


def _parse_api_date(date_str: str) -> datetime:
    normalized = date_str.replace("Z", "+00:00")
    normalized = _FRACTIONAL_SECONDS_RE.sub(lambda m: "." + m.group(1).ljust(6, "0")[:6], normalized)
    return datetime.fromisoformat(normalized)

# The spreadsheet's Date column is a local (US Eastern) date with no time
# component; the API's `date` is a UTC timestamp. Evening matches roll over
# to the next UTC calendar day (confirmed against real matches: API
# 2025-08-17T00:03:46Z == sheet 2025-08-16, consistently across many pairs at
# an offset matching America/New_York, not Pacific). Reconciliation below
# converts API timestamps to Eastern before comparing dates — comparing raw
# UTC date strings against the sheet silently missed ~14 real duplicates.
TEAM_TZ = ZoneInfo("America/New_York")

from .config import load_config
from .db import connect, init_schema, upsert
from .normalize import normalize_season_id

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("wild_tracker.xlsx_import")

WILD_TEAM_ID = "9f28fcd0-2593-49ab-9db1-170eb1d249e6"  # resolved once via premier team lookup

# Spreadsheet first-name -> (player_id/puuid, riot_name, riot_tag).
# Resolved 2026-08-16 from the user directly + HenrikDev account lookup.
# Rafid's account exists but HenrikDev/Riot couldn't return a puuid (needs a
# recent match to populate) — stored with a synthetic id, no real puuid.
PLAYER_MAP: dict[str, tuple[str, str, str]] = {
    "Calum": ("c0225528-834e-5705-8918-ad1b6b5c9c97", "YourMother", "IsAHo"),
    "Zohaib": ("cfc1950e-db2e-5bb5-a304-ae4af8dfefa5", "LegZM", "NA1"),
    "Johan": ("4e04bddb-2904-5c1c-8dca-5aad6c685412", "mochi", "LEG"),
    "Daniel": ("00160431-62ab-5d08-be2b-db7724137417", "YoungMetroDaniel", "NA1"),
    "Eli": ("4689d9a2-577b-50b7-8116-59ce5c64b490", "Smeege", "Doggo"),
    "Ei": ("4689d9a2-577b-50b7-8116-59ce5c64b490", "Smeege", "Doggo"),  # typo for Eli, per user
    "An": ("308584f8-0993-5d8c-aac6-3ab1285805a5", "Boney M", "JLee"),
    "Erik": ("2f452013-77d1-56e4-a40e-f4d719f0849b", "ColonialFungus", "BLU"),
    "Josh": ("61fc9ee9-d20a-5d64-9148-3e636121e34e", "chooper", "train"),
    "Rafid": ("xlsx-player-sultan-1479", "Sultan", "1479"),
    "Sameer": ("4a807314-981b-589d-95b5-5f88910a10ea", "krazyburro", "rito"),
    "Waqas": ("f9ed4d2c-83ba-5a25-8422-16e3e85e7c13", "losing my mind", "taco"),
}

# Data sheet column indices (1-based), from the real workbook.
COL = {
    "match_num": 1, "date": 2, "map": 3, "team_rounds": 4, "enemy_rounds": 5,
    "result": 6, "player": 8, "agent": 10, "acs": 11, "kills": 12, "deaths": 13,
    "assists": 14, "adr": 18, "hs_pct": 19, "kast_pct": 20, "fk": 21, "fd": 22, "role": 25,
    "two_k": 30, "three_k": 31, "four_k": 32, "five_k": 33,
    "1v1": 34, "1v2": 35, "1v3": 36, "1v4": 37, "1v5": 38,
    "econ": 39, "pl": 40, "de": 41, "margin": 42, "phase": 43, "type": 44,
}
WEAPON_COLS = {
    49: "Classic", 50: "Shorty", 51: "Frenzy", 52: "Ghost", 53: "Sheriff",
    54: "Stinger", 55: "Spectre", 56: "Bucky", 57: "Judge", 58: "Bulldog",
    59: "Guardian", 60: "Phantom", 61: "Vandal", 62: "Marshal", 63: "Outlaw",
    64: "Operator", 65: "Ares", 66: "Odin", 67: "Melee", 68: "Ability",
}

FIRST_UNREACHABLE_MATCH_NUM = 47  # E8A1 onward is at least sometimes API-reachable; below this, never


def load_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Data"]
    rows = []
    for row in range(2, ws.max_row + 1):
        match_num = ws.cell(row=row, column=COL["match_num"]).value
        if match_num is None:
            continue
        rows.append({key: ws.cell(row=row, column=idx).value for key, idx in COL.items()})
        rows[-1]["_weapon_kills"] = {
            name: ws.cell(row=row, column=idx).value
            for idx, name in WEAPON_COLS.items()
            if ws.cell(row=row, column=idx).value
        }
        rows[-1]["_row"] = row
    return rows


def existing_api_match_keys(conn) -> set[tuple[str, str, int]]:
    """(Eastern-local date, map, margin) keys for already-ingested API matches.

    date+map alone collides on nights with two matches on the same map
    (confirmed happens in this data) — margin (exact round differential)
    disambiguates those without needing a shared match id, which doesn't
    exist between the two sources.
    """
    cur = conn.execute("SELECT date, map, margin FROM matches WHERE source = 'api'")
    keys = set()
    for row in cur.fetchall():
        utc_dt = _parse_api_date(row["date"])
        local_date = utc_dt.astimezone(TEAM_TZ).date().isoformat()
        keys.add((local_date, row["map"], row["margin"]))
    return keys


def run_import(xlsx_path: Path) -> None:
    cfg = load_config()
    conn = connect(cfg.database_url)
    init_schema(conn)

    rows = load_rows(xlsx_path)
    by_match: dict[float, list[dict]] = {}
    for r in rows:
        by_match.setdefault(r["match_num"], []).append(r)

    skip_keys = existing_api_match_keys(conn)
    logger.info("Loaded %d player-match rows across %d matches; %d API matches to reconcile against",
                len(rows), len(by_match), len(skip_keys))

    imported, skipped_dup, skipped_unmapped = 0, 0, 0

    for match_num, player_rows in sorted(by_match.items()):
        first = player_rows[0]
        date_val = first["date"]
        date_iso = date_val.date().isoformat() if hasattr(date_val, "date") else str(date_val)
        map_name = first["map"]

        team_rounds = first["team_rounds"] or 0
        enemy_rounds = first["enemy_rounds"] or 0
        # A few rows have a blank Margin formula cell in the source file
        # (confirmed: match #85) despite having rounds-won data — recompute
        # rather than leaving the dedup key (and the stored margin) null.
        margin_val = first["margin"] if first["margin"] is not None else (team_rounds - enemy_rounds)
        dup_key = (date_iso, map_name, int(margin_val))
        if match_num >= FIRST_UNREACHABLE_MATCH_NUM and dup_key in skip_keys:
            skipped_dup += 1
            continue

        match_id = f"xlsx-{int(match_num):03d}"

        upsert(conn, "matches", {
            "match_id": match_id,
            "source": "spreadsheet",
            "season_id": normalize_season_id(first["phase"]),
            "date": date_iso,
            "map": map_name,
            "match_type": first["type"],
            "team_id": WILD_TEAM_ID,
            "enemy_team_id": None,
            "result": first["result"],
            "margin": margin_val,
            "raw_payload": None,
        })

        for r in player_rows:
            name = r["player"]
            if name not in PLAYER_MAP:
                logger.warning("Match %s: unmapped player name %r, skipping this player row", match_num, name)
                skipped_unmapped += 1
                continue
            player_id, riot_name, riot_tag = PLAYER_MAP[name]

            upsert(conn, "players", {"player_id": player_id, "riot_name": riot_name, "riot_tag": riot_tag})

            def as_pct(val):
                return val * 100 if isinstance(val, (int, float)) and val <= 1 else val

            upsert(conn, "match_players", {
                "match_id": match_id,
                "player_id": player_id,
                "team_id": WILD_TEAM_ID,
                "agent": r["agent"],
                "role": r["role"],
                "score": None,
                "kills": r["kills"],
                "deaths": r["deaths"],
                "assists": r["assists"],
                "headshots": None,
                "bodyshots": None,
                "legshots": None,
                "damage_dealt": None,
                "damage_received": None,
                "economy_spent_overall": None,
                "economy_loadout_value_overall": None,
                "rounds_played": team_rounds + enemy_rounds,
                "adr": r["adr"],
                "hs_pct": as_pct(r["hs_pct"]),
                "fk": r["fk"],
                "fd": r["fd"],
                "acs": r["acs"],
                "kast_pct": as_pct(r["kast_pct"]),
            })

            # A blank spreadsheet cell means "zero of this," not "unknown" —
            # store 0 rather than NULL, otherwise SQL's SUM(a+b+c+d+e) style
            # aggregates (e.g. queries.py's clutch totals) silently drop the
            # whole row whenever any one of the five clutch columns is NULL,
            # undercounting spreadsheet-era stats (found 2026-08-31: metro's
            # career clutch total was reading 12, not the real ~34+, because
            # nearly every spreadsheet row had 4 of 5 clutch columns NULL).
            upsert(conn, "derived_player_match_stats", {
                "match_id": match_id,
                "player_id": player_id,
                "source": "spreadsheet_manual",
                "two_k": r["two_k"] or 0,
                "three_k": r["three_k"] or 0,
                "four_k": r["four_k"] or 0,
                "five_k": r["five_k"] or 0,
                "clutch_1v1": r["1v1"] or 0,
                "clutch_1v2": r["1v2"] or 0,
                "clutch_1v3": r["1v3"] or 0,
                "clutch_1v4": r["1v4"] or 0,
                "clutch_1v5": r["1v5"] or 0,
                "plants": r["pl"] or 0,
                "defuses": r["de"] or 0,
                "econ": r["econ"],
            })

            for weapon, count in r["_weapon_kills"].items():
                upsert(conn, "match_player_weapon_kills", {
                    "match_id": match_id, "player_id": player_id, "weapon": weapon, "kill_count": count,
                })

        imported += 1

    conn.commit()
    logger.info(
        "Import complete: %d matches imported, %d skipped as API duplicates, %d unmapped player rows skipped",
        imported, skipped_dup, skipped_unmapped,
    )
    conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Import legacy matches from the WILD Excel workbook.")
    parser.add_argument("xlsx_path", type=Path)
    args = parser.parse_args()
    run_import(args.xlsx_path)


if __name__ == "__main__":
    main()
